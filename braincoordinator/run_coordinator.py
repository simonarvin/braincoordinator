import sys
from os import system, name, path
from os.path import dirname

import cv2

import numpy as np
from openpyxl import load_workbook
from shutil import copyfile

from datetime import datetime

from braincoordinator.utilities.manager import Manager
from braincoordinator.utilities.computations import *
from braincoordinator.utilities.arguments import Arguments
from braincoordinator.utilities.get_atlas import Getter

font = cv2.FONT_HERSHEY_SIMPLEX

class Coordinator:
    def __init__(self, args, ap:float = 0, ml:float = 0, dv:float =0) -> None:

        self.arguments = Arguments(args)

        if self.get(self.arguments.get):
            return

        self.reference = int(str(self.arguments.reference[0]).lower() == "l") #bregma = 0; lambda = 1
        self.counterreference = int(self.reference == 0)
        self.dir_path = dirname(path.realpath(__file__))
        self.animal = self.arguments.animal
        self.preload = self.arguments.preload

        animal_path = "{}/atlas/{}/".format(self.dir_path, self.animal)

        try:
            self.manager = Manager(animal_path, self.preload, self.reference)
        except FileNotFoundError as e:
            print(e)
            print("")
            print("Atlas not available.")
            print("Write '--get [atlas]' to download an atlas.")
            print("Or, write '--get list' to list all atlases.")
            return


        self.selected_marker = None

        self.markers = []
        self.paths = []

        self.primary_color = [0, 0, 255]
        self.second_color = [50, 50, 255]

        self.x, self.y = [0,0], [0,0]
        self.hover_window = 0

        self.print_instructions()

        self.manager.set_values(ap, ml, dv)
        self.iterate()

    def get(self, get:str):

        if get != "":
            Getter(get)
            return True

        return False

    def print_instructions(self):
        instructions = """
                        BRAIN COORDINATOR
            Developed by Simon Arvin
            https://github.com/simonarvin/braincoordinator

        INSTRUCTIONS:
        A/S - previous/next coronal slice.
        Z/X - previous/next sagittal slice.
        D   - place marker.
        F   - remove marker.
        P   - save data-file.

        Click and drag to move markers.
        """

        print(instructions)

    def frontal_mouse(self, event, x:float, y:float, flags, param) -> None:

        if event == cv2.EVENT_MOUSEMOVE:
            self.hover_window = 0

            if self.selected_marker != None:
                if self.selected_marker[3] == self.hover_window:
                    self.manager.update_marker(self.selected_marker, (x, y), self.hover_window)
                    coronal_image, sagittal_image = self.update()
                    cv2.imshow("Coronal", coronal_image)
                    cv2.imshow("Sagittal", sagittal_image)


        if event == cv2.EVENT_LBUTTONDOWN:
            if self.selected_marker == None:
                distance, marker = self.find_nearest_marker((x, y))
                if distance < 20:
                    self.selected_marker = marker

        elif event == cv2.EVENT_LBUTTONUP:

            self.selected_marker = None

        self.x[0], self.y[0] = x, y

    def sagittal_mouse(self, event, x:float, y:float, flags, param) -> None:

        if event == cv2.EVENT_MOUSEMOVE:
            self.hover_window = 1

            if self.selected_marker != None:
                if self.selected_marker[3] == self.hover_window:
                    self.manager.update_marker(self.selected_marker, (x, y), self.hover_window)
                    coronal_image, sagittal_image = self.update()
                    cv2.imshow("Coronal", coronal_image)
                    cv2.imshow("Sagittal", sagittal_image)

        if event == cv2.EVENT_LBUTTONDOWN:
            if self.selected_marker == None:
                distance, marker = self.find_nearest_marker((x, y))
                if distance < 20:
                    self.selected_marker = marker

        elif event == cv2.EVENT_LBUTTONUP:

            self.selected_marker = None

        self.x[1], self.y[1] = x, y

    def place_cross(self, source: np.ndarray, point: tuple, color: tuple) -> None:
        source[point[1] - 7:point[1] + 8, point[0]] = color
        source[point[1], point[0] - 7:point[0] + 8] = color
        #cv2.circle(source, point,4,color,1)



    def get_scale(self, img, type:int = 0):
        """
        This function is used to extract scales for all slices. We save this data as a .sc file and then use set_scale instead.
        type 0 => horisontal axis
        type 1 => vertical axis
        """
        i = 0
        if type == 0:
            width = int(img.shape[1]/2) + 50
            offset = 45
        else:
            height = int(img.shape[0]/4) + 5
            offset = 45

        start = 0
        ready = 0

        try:
            while i < 500:
                if type == 0:
                    pixel = img[offset, width + i]
                else:
                    pixel = img[height + i, offset]

                if pixel[0] < 245:
                    if start == 0:
                        if type == 0:
                            start = width+i
                        else:
                            start = height + i

                    elif ready > 5:
                        if type == 0:
                            cv2.line(img, (start,offset), (width+i,offset), self.primary_color, 1)
                            return [start, abs(start - (width + i))] #center, pixels per mm
                        else:
                            cv2.line(img, (offset, start), (offset, height+i), self.primary_color, 1)
                            return [start, abs(start - (height + i))] #center, pixels per mm

                elif start != 0:

                    ready += 1
                i += 1
        except:
            pass # out of bounds


    def find_nearest_marker(self, mouse_pos:tuple) -> tuple:

        distance_pairs = []

        for marker in self.markers:
            if marker[3] == self.hover_window:
                if (marker[3] == 0 and self.manager.coronal_index == marker[1]) or (marker[3] == 1 and self.manager.sagittal_index == marker[1]):
                    distance_pairs.append([distance2d(marker[0], mouse_pos), marker])

        try:
            distance_pairs = sorted(distance_pairs, key=lambda x: float(x[0]))
            return distance_pairs[0]
        except:
            return (30, -1)


    def update(self) -> tuple:

        coronal_image, sagittal_image = self.manager.get_images()

        #self.scale = self.get_scale(img)  #tuple center, pixels per mm
        #self.scale[0] = self.scale[0] - self.scale[1] * 2
        #img[self.scale[0],30] = self.primary_color
        #print(self.scale)

        self.manager.set_scale()

        #self.clear()
        self.print_instructions()
        print("Markers")

        for i, marker in enumerate(self.markers):

            print("     M{} - ap: {}; ml: {}; dv: {}".format(i,marker[2][0],marker[2][1],marker[2][2]))

            if (i + 1) % 2 == 0:
                angle_front = np.degrees(np.arctan2(-self.markers[i - 1][2][1] + marker[2][1], -self.markers[i-1][2][2] + marker[2][2]))
                angle_sag = np.degrees(np.arctan2(-self.markers[i - 1][2][0]+marker[2][0], -self.markers[i-1][2][2] + marker[2][2]))
                distance = distance3d(marker[2], self.markers[i - 1][2])

                if len(self.paths) > 0:
                    if i != self.paths[len(self.paths)-1][0]:
                        self.paths.append([i, self.markers[i - 1], marker, angle_front, angle_sag, distance])
                else:
                    self.paths.append([i,self.markers[i - 1], marker, angle_front, angle_sag, distance])


                print("     |M{}M{}| {} mm; ang_cor: {} deg; ang_sag: {} deg".format(i - 1, i, round(distance, 2), np.round(angle_front, 2), np.round(angle_sag, 2)))


            if marker[3] == 0: #frontal
                if marker[1] == self.manager.coronal_index:

                    self.place_cross(coronal_image, marker[0],self.primary_color)

                    if (i + 1) % 2 == 0:
                        cv2.line(coronal_image, self.markers[i - 1][0], marker[0], self.primary_color, 1)

                    cv2.putText(coronal_image, "M"+str(i), tuple([mark + 5 for mark in marker[0]]), font,  .5, self.primary_color, 1, cv2.LINE_AA)

                coord_float=str_to_float(self.manager.coordinate[1])
                if (i + 1) % 2 == 0:

                    if self.markers[i - 1][2][1] > coord_float > marker[2][1] or self.markers[i - 1][2][1] < coord_float < marker[2][1]:
                        fraction = (coord_float - self.markers[i - 1][2][1])/(marker[2][1] - self.markers[i - 1][2][1])

                        new_marker = self.manager.to_pixel(marker[2], 1)
                        old_marker = self.manager.to_pixel(self.markers[i - 1][2], 1)

                        ml_diff = (new_marker[0] - old_marker[0]) * fraction
                        dv_diff = (new_marker[1] - old_marker[1]) * fraction

                        cv2.line(sagittal_image, old_marker, new_marker, self.second_color, 1)
                        cv2.circle(sagittal_image, old_marker, 2, self.second_color, -1)
                        cv2.circle(sagittal_image, new_marker, 2, self.second_color, -1)

                        start = np.array([old_marker[0] + ml_diff, old_marker[1] + dv_diff], dtype = int)

                        cv2.circle(sagittal_image, tuple(start), 4, self.primary_color, -1)

                size = max(.7 - abs(coord_float - marker[2][1]) * .2, .3)
                new_marker = self.manager.to_pixel(marker[2], 1)
                self.place_cross(sagittal_image, new_marker, self.primary_color)
                cv2.putText(sagittal_image, "M" + str(i), tuple([mark + 5 for mark in new_marker]), font,  size, self.primary_color, 1, cv2.LINE_AA)

            else:

                if marker[1] == self.manager.sagittal_index:
                    self.place_cross(sagittal_image, marker[0], self.primary_color)

                    if (i + 1) % 2 == 0:
                        cv2.line(sagittal_image, self.markers[i - 1][0], marker[0], self.primary_color, 1)

                    cv2.putText(sagittal_image, "M" + str(i), tuple([mark + 5 for mark in marker[0]]), font, .5, self.primary_color, 1, cv2.LINE_AA)

                coord_float = str_to_float(self.manager.coordinate[0])
                if (i + 1) % 2 == 0:

                    if self.markers[i - 1][2][0] > coord_float > marker[2][0] or self.markers[i - 1][2][0] < coord_float < marker[2][0]:

                        fraction = (coord_float - self.markers[i - 1][2][0])/(marker[2][0] - self.markers[i - 1][2][0])

                        new_marker = self.manager.to_pixel(marker[2], 0)
                        old_marker = self.manager.to_pixel(self.markers[i - 1][2], 0)
                        ml_diff = (new_marker[0] - old_marker[0]) * fraction
                        dv_diff = (new_marker[1] - old_marker[1]) * fraction

                        cv2.line(coronal_image, old_marker, new_marker, self.second_color, 1)
                        cv2.circle(coronal_image, old_marker, 2, self.second_color, -1)
                        cv2.circle(coronal_image, new_marker, 2, self.second_color, -1)

                        start = np.array([old_marker[0] + ml_diff, old_marker[1] + dv_diff], dtype = int)

                        cv2.circle(coronal_image, tuple(start), 4, self.primary_color, -1)

                size = max(.7 - abs(coord_float - marker[2][0]) * .2, .3)
                new_marker = self.manager.to_pixel(marker[2], 0)
                self.place_cross(coronal_image, new_marker, self.primary_color)
                cv2.putText(coronal_image, "M" + str(i), tuple([mark + 5 for mark in new_marker]), font,  size, self.primary_color, 1, cv2.LINE_AA)

        return coronal_image, sagittal_image


    def clear(self):

        # for windows
        if name == 'nt':
            _ = system('cls')

        # for mac and linux(here, os.name is 'posix')
        else:
            _ = system('clear')

    def copy_paste_xcl(self, startCol, startRow, endCol, endRow, sheet, data_out):

        rangeSelected = []
        #Loops through selected Rows
        for i in range(startRow, endRow + 1,1):
        #Appends the row to a RowSelected list
            rowSelected = []
            for j in range(startCol,endCol+1,1):
                rowSelected.append(sheet.cell(row = i, column = j).value)
            #Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)

        q=0
        for path in data_out:

            #n+=1

            countRow = 0

            #todo: add one marker also

            rangeSelected[5][1]=path[3] #front ang
            rangeSelected[6][1]=path[4] #sag ang
            rangeSelected[7][1]=path[5] #distance

            rangeSelected[13][1]="=B{}-180/PI()*MOD(G3,PI())".format(12+17*q)
            rangeSelected[14][1]="=B{}-180/PI()*MOD(G4,PI())".format(13+17*q)
            rangeSelected[15][1]="=((B{}-B{})^2+(C{}-C{})^2+(D{}-D{})^2)^(1/2)".format(18+17*q,19+17*q,18+17 *q,19+17*q,18+17*q,19+17*q)

            #Marker 1

            rangeSelected[3][0]="m{}".format(path[0]-1)
            rangeSelected[3][1]="=B3 + {}".format(path[1][2][0])#ap
            rangeSelected[3][2]="=C3 + {}".format(path[1][2][1])#ml
            rangeSelected[3][3]="=D3 + {}".format(path[1][2][2])#dv

            #Marker 2

            rangeSelected[4][0]="m{}".format(path[0])
            rangeSelected[4][1]="=B3 + {}".format(path[2][2][0])#ap
            rangeSelected[4][2]="=C3 + {}".format(path[2][2][1])#ml
            rangeSelected[4][3]="=D3 + {}".format(path[2][2][2])#dv

            #AP
            rangeSelected[11][1]="=B3 + G7 * (B{} - B3) - H7 * (C{}-C3)".format(10 + 17*q, 10+17*q)
            rangeSelected[12][1]="=B3 + G7 * (B{} - B3) - H7 * (C{}-C3)".format(11 + 17 *q, 11+17*q)


            rangeSelected[11][0]="m{}".format(path[0] -1)
            rangeSelected[12][0]="m{}".format(path[0])
            #ML
            rangeSelected[11][2]="=H7 * (B{} - B3) +G7 * (C{} - C3)".format(10 + 17 * q, 10 + 17 * q)
            rangeSelected[12][2]="=H7 * (B{} - B3) +G7 * (C{} - C3)".format(11 + 17 * q, 11 + 17 * q)

            #DV
            rangeSelected[11][3]="=D{}".format(10 + 17 * q)
            rangeSelected[12][3]="=D{}".format(11 + 17 * q)

            q+=1

            for i in range(startRow, endRow+1,1):
                countCol = 0
                for j in range(startCol, endCol+1,1):

                    sheet.cell(row = i, column = j).value = rangeSelected[countRow][countCol]
                    countCol += 1
                countRow += 1
            startRow += 17
            endRow += 17


    def save_data(self):
        if len(self.paths) == 0:
            print("No paths marked.")
            return

        now = datetime.now()
        time = now.strftime("%Y%m%d%H%M%S")
        new_save = self.dir_path+"/data/{}_coordinates_{}.xlsx".format(self.animal, time)
        copyfile(self.dir_path+"/data/do-not-delete.xlsx", new_save)
        workbook = load_workbook(filename=new_save)
        sheet = workbook.active
        self.copy_paste_xcl(1, 7, 4, 22, sheet, self.paths)
        workbook.save(filename = new_save)
        print("Data saved")

    def keyHandler(self, key):
        if key == ord("x"):
            self.manager.next("sagittal")

        elif key == ord("z"):
            self.manager.previous("sagittal")

        elif key == ord("a"):
            self.manager.next("coronal")

        elif key == ord("s"):
            self.manager.previous("coronal")

        elif key==ord("d"):
            try:
                point = (self.x[self.hover_window],self.y[self.hover_window])
                if self.hover_window==0:
                    self.markers.append([point, self.manager.coronal_index, self.manager.convert_to_mm(point, 0), self.hover_window])
                else:
                    self.markers.append([point, self.manager.sagittal_index, self.manager.convert_to_mm(point, 1), self.hover_window])
            except AttributeError:
                print("Click image once to gain focus.")

        elif key == ord("f"):
            try:
                if len(self.markers)%2 == 0:
                    self.paths.pop()

                self.markers.pop()
            except:
                pass
        elif key == ord("p"):
            self.save_data()
        elif key == ord("q"):
            return True

        return False

    def iterate(self):

        #todo: du mangler angles og angle korrektion i save-file

        self.manager.coronal_index, self.manager.sagittal_index = self.manager.find_nearest_slices()

        coronal_image, sagittal_image = self.update()

        cv2.imshow("Coronal", coronal_image)
        cv2.imshow("Sagittal", sagittal_image)
        cv2.setMouseCallback("Coronal", self.frontal_mouse)
        cv2.setMouseCallback("Sagittal", self.sagittal_mouse)

        while True:

            coronal_image, sagittal_image = self.update()

            cv2.imshow("Coronal", coronal_image)
            cv2.imshow("Sagittal", sagittal_image)

            key = cv2.waitKey(0)

            if self.keyHandler(key):
                break

def main():
    print("Brain coordinator initiated")
    Coordinator(sys.argv[1:])

if __name__ == '__main__':
    main()
